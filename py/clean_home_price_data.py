import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
from openpyxl import load_workbook


path = '/Users/nathanoliver/Desktop/Python/Home_Prices_Engineer_Salaries/xlsx/me_salaries.xlsx'

print('loading workbook')
wb = load_workbook(path)

sheet2 = wb["Sheet2"]
sheet3 = wb["Sheet3"]


k = sheet2.cell(row=1, column=1).value


info = k.split()

# print(info)

cell = info[0]

col_names = ['cbsa_code', '2018', '2019', '2020', '2020.II',
             '2020.III', '2020.IV', '2021.Ir', '2021.IIp', 'Q2-Q2']

for i in range(10):
    sheet3.cell(row=1, column=i + 1).value = col_names[i]

col = 0
row = 1

for i in range(len(info)):
    cell = info[i]
    if col == 0:
        print('entered zero column')
        row = row + 1
        col = col + 1
        sheet3.cell(row=row, column=col).value = info[i]
    elif col > 0:
        print('entered non-zero column')
        if cell[0].isdigit() == True:
            col = col + 1
            sheet3.cell(row=row, column=col).value = info[i]
        # if cell[0].isdigit() == False:
    if col == 10:
        print('entered change row')
        col = 0


wb.save(path)
